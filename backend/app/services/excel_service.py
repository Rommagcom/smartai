"""Excel file generation service using openpyxl."""
from __future__ import annotations

import base64
import logging
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


class ExcelService:
    """Generate .xlsx files from structured or plain-text data."""

    def create_excel_base64(
        self,
        title: str,
        content: str,
        filename: str = "document.xlsx",
        columns: list[str] | None = None,
        rows: list[list] | None = None,
    ) -> dict:
        """Create an Excel file and return base64-encoded result.

        If *rows* (and optionally *columns*) are provided, they are used
        directly as structured data.  Otherwise *content* is parsed as
        tab/semicolon-separated text where the first line is the header.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31] if title else "Sheet1"  # Excel limits sheet name to 31 chars

        if rows:
            self._fill_structured(ws, columns, rows)
        else:
            self._fill_from_text(ws, content)

        self._auto_column_width(ws)

        buf = BytesIO()
        wb.save(buf)
        blob = buf.getvalue()

        return {
            "file_name": filename,
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "file_base64": base64.b64encode(blob).decode("utf-8"),
            "size_bytes": len(blob),
        }

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _fill_structured(
        self,
        ws,
        columns: list[str] | None,
        rows: list[list],
    ) -> None:
        start_row = 1
        if columns:
            for col_idx, header in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGNMENT
            start_row = 2

        for row_idx, row_data in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = _CELL_ALIGNMENT

    def _fill_from_text(self, ws, content: str) -> None:
        """Parse plain text (TSV / semicolon / CSV) into rows."""
        lines = [l for l in content.splitlines() if l.strip()]
        if not lines:
            ws.cell(row=1, column=1, value=content)
            return

        sep = self._detect_separator(lines[0])

        for row_idx, line in enumerate(lines, start=1):
            values = [v.strip() for v in line.split(sep)]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._coerce(value))
                if row_idx == 1:
                    cell.font = _HEADER_FONT
                    cell.fill = _HEADER_FILL
                    cell.alignment = _HEADER_ALIGNMENT
                else:
                    cell.alignment = _CELL_ALIGNMENT

    @staticmethod
    def _detect_separator(line: str) -> str:
        """Pick the most likely column separator."""
        for sep in ("\t", ";", "|"):
            if sep in line:
                return sep
        if line.count(",") >= 2:
            return ","
        return "\t"

    @staticmethod
    def _coerce(value: str):
        """Try to convert string value to int/float for proper Excel typing."""
        if not value:
            return value
        # integer
        if re.fullmatch(r"-?\d+", value):
            try:
                return int(value)
            except (ValueError, OverflowError):
                pass
        # float
        if re.fullmatch(r"-?\d+[.,]\d+", value):
            try:
                return float(value.replace(",", "."))
            except (ValueError, OverflowError):
                pass
        return value

    @staticmethod
    def _auto_column_width(ws, max_width: int = 60) -> None:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            adjusted = min(max_len + 3, max_width)
            ws.column_dimensions[col_letter].width = max(adjusted, 10)


excel_service = ExcelService()
